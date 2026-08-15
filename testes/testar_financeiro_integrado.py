"""Testes do HISTORICO e da INTEGRACAO Financeiro/Fluxo de Caixa (Fase 4).

O que estes testes protegem:

  1. VENCIMENTO ANDA COM A PARCELA. Compra em 6x nao pode nascer com seis
     vencimentos no mesmo dia — o Financeiro ordena e soma por vencimento, entao
     isso empilhava seis meses de despesa num mes so.
  2. DIA 31 EM MES CURTO. Cai no ultimo dia valido e VOLTA pro 31 no mes
     seguinte. Encadear a partir da parcela anterior faria a data escorregar.
  3. SIMULADOR NAO ESCREVE. Nem uma linha.
  4. HISTORICO PROTEGIDO. Parcela paga, conciliada ou com PIX nao e reescrita
     nem quando esta marcada.
  5. UMA FONTE SO. Financeiro e Fluxo de Caixa leem os mesmos numeros.
  6. EXPORTACAO HONESTA. O arquivo declara qual recorte ele e.

Como rodar:

    JOB_MODO_TESTE=1 JOB_DATA_DIR=/tmp/jobtest python3 testes/testar_financeiro_integrado.py
"""
import os
import sys
import json
import hashlib
import tempfile

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.environ.setdefault('JOB_MODO_TESTE', '1')
os.environ.setdefault('JOB_DATA_DIR', tempfile.mkdtemp(prefix='jobtest-fin-'))

import app as A  # noqa: E402

falhas = []


def checa(nome, cond, detalhe=''):
    if cond:
        print(f'  ok   {nome}')
    else:
        print(f'  FALHA {nome} {detalhe}')
        falhas.append(f'{nome} {detalhe}')


def cliente(perfil='admin'):
    c = A.app.test_client()
    with c.session_transaction() as s:
        s['user_id'] = 1
        s['perfil'] = perfil
        s['nome'] = 'Gestor Teste'
    return c


def limpar():
    conn = A.db()
    for t in ('proposta_regra_snapshot', 'gestor_retencao', 'gestor_regra',
              'affinity_conciliacao', 'fin_evento'):
        conn.execute(f"DELETE FROM {t}")
    conn.execute("DELETE FROM lancamentos WHERE descricao LIKE 'TESTE FIN%'")
    conn.execute("DELETE FROM parcelas WHERE proposta_id IN "
                 "(SELECT id FROM propostas WHERE razao_social LIKE 'TESTE FIN%')")
    conn.execute("DELETE FROM propostas WHERE razao_social LIKE 'TESTE FIN%'")
    conn.execute("DELETE FROM usuarios WHERE email LIKE 'teste.fin%'")
    conn.commit()
    A.close_db(conn)


def foto_do_banco():
    conn = A.db()
    tabelas = [r[0] for r in conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' ORDER BY name").fetchall()]
    foto = {}
    for t in tabelas:
        if t.startswith('sqlite_'):
            continue
        h = hashlib.sha256()
        linhas = conn.execute(f'SELECT * FROM "{t}"').fetchall()
        for r in linhas:
            h.update(repr(tuple(r)).encode('utf-8', 'replace'))
        foto[t] = (len(linhas), h.hexdigest())
    A.close_db(conn)
    return foto


def montar_venda_completa(operadora='Amil', pago=False, pix=False, conciliada=False):
    """Venda de gestor com regra completa. Opcionalmente ja com dinheiro andado."""
    conn = A.db()
    cur = conn.execute("""INSERT INTO usuarios (nome, email, perfil, regime_base, ativo)
                          VALUES (?,?,?,'',1)""",
                       ('Gestor Fin', f'teste.fin.{operadora.lower()}@x.com', 'admin'))
    uid = A._last_insert_id(cur)
    cur = conn.execute("""INSERT INTO gestor_regra
        (operadora, obs, plano, fracoes_json, gestor_json, confirmada, ativo, criado_por, criado_em)
        VALUES (?,'','PME',?,?,1,1,'teste',?)""",
        (operadora,
         json.dumps([{'ordem': 1, 'percentual': 30.0, 'evento': '1a mensalidade', 'mes': 1},
                     {'ordem': 2, 'percentual': 70.0, 'evento': '2a mensalidade', 'mes': 2}]),
         json.dumps([{'ordem': 1, 'percentual_gestor': 100.0},
                     {'ordem': 2, 'percentual_gestor': 0.0}]), A._agora_sp()))
    rid = A._last_insert_id(cur)
    conn.execute("""INSERT INTO gestor_retencao
        (regra_id, tipo, nome, percentual, base_calculo, responsavel, ativo, criado_por, criado_em)
        VALUES (?,'imposto','ISS',10.0,'bruto_gestor','gestor',1,'teste',?)""",
        (rid, A._agora_sp()))
    # regime_aplicado e obrigatorio pra venda entrar no motor do gestor: _fin_visao
    # so soma quem esta em socio_gestor_regra/socio_gestor_pendente. Sem isso o
    # fixture montava a venda e ela ficava invisivel pra propria conta que o teste
    # queria verificar.
    cur = conn.execute("""INSERT INTO propostas (usuario_id, consultor, numero_proposta,
                    razao_social, status, comissao_total_corretora, comissao_consultor,
                    vigencia, modalidade, tipo_contrato, acomodacao, fator_moderador,
                    total_vidas, valor, adm_operadora, regime_aplicado, criado_em)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                 (uid, 'Gestor Fin', '88800001', f'TESTE FIN {operadora} LTDA', 'Implantada',
                  1000.0, 200.0, '2026-07-01', 'PME', 'Novo', 'Enfermaria', 'Sem', 1, 1000.0,
                  operadora, 'socio_gestor_regra', A._agora_sp()))
    pid = A._last_insert_id(cur)
    cur = conn.execute("""INSERT INTO parcelas (proposta_id, numero, percentual, valor, status,
                          asaas_transfer_id) VALUES (?,?,?,?,?,?)""",
                       (pid, 1, 100, 300.0,
                        'Pago ao corretor' if pago else 'Pendente de receber',
                        'tra_999' if pix else None))
    conn.commit()
    A._gestor_congelar_snapshot(conn, pid, operadora, 'PME', uid, 'Gestor Fin')
    if conciliada:
        conn.execute("""INSERT INTO affinity_conciliacao
            (chave_idem, codigo_comissao, arquivo, proposta_id, liquido, estado, criado_em)
            VALUES (?,?,?,?,?,'entrada_confirmada',?)""",
            (f'afy:teste{operadora}:0', '9000001', 'x.pdf', pid, 300.0, A._agora_sp()))
    conn.commit()
    A.close_db(conn)
    return pid, uid


def teste_vencimento_por_parcela():
    print('\n[unit] vencimento avanca um mes por parcela')
    checa('parcela 1 mantem a data', A._venc_parcela('2026-03-15', 0) == '2026-03-15')
    checa('parcela 2 vai pro mes seguinte', A._venc_parcela('2026-03-15', 1) == '2026-04-15',
          A._venc_parcela('2026-03-15', 1))
    checa('parcela 6 anda cinco meses', A._venc_parcela('2026-03-15', 5) == '2026-08-15',
          A._venc_parcela('2026-03-15', 5))
    checa('sem data continua sem data', A._venc_parcela('', 3) == '')
    checa('data invalida nao estoura', A._venc_parcela('xx', 2) == 'xx')


def teste_politica_dia_31():
    print('\n[unit] dia 29, 30 e 31 em mes curto')
    # 2026 nao e bissexto: fevereiro tem 28.
    checa('31/01 vira 28/02', A._venc_parcela('2026-01-31', 1) == '2026-02-28',
          A._venc_parcela('2026-01-31', 1))
    checa('31/01 VOLTA pra 31/03 (nao fica preso no 28)',
          A._venc_parcela('2026-01-31', 2) == '2026-03-31',
          A._venc_parcela('2026-01-31', 2))
    checa('31/01 vira 30/04', A._venc_parcela('2026-01-31', 3) == '2026-04-30',
          A._venc_parcela('2026-01-31', 3))
    checa('31/01 volta pra 31/05', A._venc_parcela('2026-01-31', 4) == '2026-05-31',
          A._venc_parcela('2026-01-31', 4))
    checa('30/01 vira 28/02', A._venc_parcela('2026-01-30', 1) == '2026-02-28',
          A._venc_parcela('2026-01-30', 1))
    checa('29/01 vira 28/02 em ano comum', A._venc_parcela('2026-01-29', 1) == '2026-02-28',
          A._venc_parcela('2026-01-29', 1))
    # 2028 e bissexto: fevereiro tem 29.
    checa('29/01 vira 29/02 em ano bissexto', A._venc_parcela('2028-01-29', 1) == '2028-02-29',
          A._venc_parcela('2028-01-29', 1))
    checa('31/12 vira 31/01 do ano seguinte', A._venc_parcela('2026-12-31', 1) == '2027-01-31',
          A._venc_parcela('2026-12-31', 1))


def teste_lancamento_parcelado_na_rota():
    print('\n[unit] lancamento em 6x pela rota')
    limpar()
    c = cliente()
    r = c.post('/lancamento/salvar', json={
        'tipo': 'custo', 'descricao': 'TESTE FIN COMPRA PARCELADA', 'valor': 600.0,
        'data_emissao': '2026-01-31', 'data_vencimento': '2026-01-31', 'num_parcelas': 6,
        'centro_custo': 'Mídia'})
    checa('rota responde', r.status_code == 200 and r.get_json().get('ok'), r.status_code)
    conn = A.db()
    linhas = conn.execute("""SELECT parcela_num, data_vencimento, data_competencia, valor
                             FROM lancamentos WHERE descricao LIKE 'TESTE FIN COMPRA%'
                             ORDER BY parcela_num""").fetchall()
    A.close_db(conn)
    vencs = [l['data_vencimento'] for l in linhas]
    checa('criou 6 parcelas', len(linhas) == 6, len(linhas))
    checa('cada parcela tem vencimento diferente', len(set(vencs)) == 6, vencs)
    checa('a sequencia respeita a politica do dia 31',
          vencs == ['2026-01-31', '2026-02-28', '2026-03-31', '2026-04-30',
                    '2026-05-31', '2026-06-30'], vencs)
    checa('a soma continua sendo o valor total',
          abs(sum(float(l['valor']) for l in linhas) - 600.0) < 0.01,
          sum(float(l['valor']) for l in linhas))
    checa('competencia continua andando junto',
          [l['data_competencia'] for l in linhas][:3] == ['2026-01', '2026-02', '2026-03'],
          [l['data_competencia'] for l in linhas][:3])


def teste_simulador_nao_escreve():
    print('\n[int] simulador nao grava nada')
    limpar()
    montar_venda_completa('Amil')
    antes = foto_do_banco()
    c = cliente()
    r = c.get('/comissoes/regra-gestor/simular')
    depois = foto_do_banco()
    difs = [t for t in set(antes) | set(depois) if antes.get(t) != depois.get(t)]
    checa('simulador abre', r.status_code == 200, r.status_code)
    checa('nenhuma tabela mudou', not difs, difs)
    corpo = r.get_data(as_text=True)
    checa('a tela diz que nao muda nada sozinha', 'não muda nada sozinha' in corpo)
    # Checa o SENTIDO (as duas colunas existem), nao a redacao exata — senao
    # qualquer melhoria de texto quebra o teste sem nada ter quebrado de fato.
    baixo = corpo.lower()
    checa('mostra atual e simulado lado a lado',
          'atual' in baixo and ('simula' in baixo),
          [s for s in ('atual', 'simula') if s not in baixo])
    checa('mostra a diferenca', 'Diferença' in corpo)


def teste_historico_protegido():
    print('\n[int] parcela paga, conciliada ou com PIX nao e reescrita')
    limpar()
    pid_pago, _ = montar_venda_completa('OperadoraPaga', pago=True)
    pid_pix, _ = montar_venda_completa('OperadoraPix', pix=True)
    pid_conc, _ = montar_venda_completa('OperadoraConc', conciliada=True)
    pid_livre, _ = montar_venda_completa('OperadoraLivre')

    conn = A.db()
    for pid, rotulo in [(pid_pago, 'paga'), (pid_pix, 'com PIX'), (pid_conc, 'conciliada')]:
        travas = A._historico_travas(conn, pid)
        checa(f'venda {rotulo} esta travada', len(travas) > 0, travas)
    checa('venda sem movimento nao esta travada',
          A._historico_travas(conn, pid_livre) == [], A._historico_travas(conn, pid_livre))
    A.close_db(conn)

    c = cliente()
    # A aplicacao automatica no historico foi BLOQUEADA de proposito
    # (aplicacao_historica_bloqueada em app.py): nao se troca um repasse de
    # consultor ja cadastrado por uma regra nova sem migracao auditada. A rota
    # responde 409 e nao escreve nada. Este teste guarda esse contrato — se
    # alguem reabrir a rota sem migracao, ele quebra e avisa.
    r = c.post('/comissoes/regra-gestor/aplicar-historico',
               json={'propostas': [pid_pago, pid_pix, pid_conc, pid_livre],
                     'confirmacao': 'APLICAR'})
    d = r.get_json() or {}
    checa('aplicacao em massa no historico segue bloqueada', r.status_code == 409, r.status_code)
    checa('e diz o motivo em vez de so recusar',
          'bloqueada' in (d.get('erro') or '').lower(), d.get('erro'))
    checa('nao aplicou nada', not d.get('ok'), d)

    # Com a rota bloqueada, o certo e o oposto do teste antigo: nada pode ter
    # sido escrito. Se um dia a aplicacao for reaberta com migracao auditada,
    # este teste muda junto — e a mudanca fica explicita no diff.
    conn = A.db()
    h = conn.execute("""SELECT * FROM historico_proposta
                        WHERE proposta_id=? AND campo LIKE 'Regra do gestor%'""",
                     (pid_livre,)).fetchone()
    A.close_db(conn)
    checa('bloqueada, nao escreveu historico nenhum', h is None, h and dict(h))


def teste_uma_fonte_para_as_duas_telas():
    print('\n[int] Financeiro e Fluxo de Caixa leem a mesma fonte')
    limpar()
    pid, _ = montar_venda_completa('Amil')
    conn = A.db()
    visao = A._fin_visao(conn)
    A.close_db(conn)
    # Comissao 1000: fracao 1 = 300 (100% do gestor, ISS 10% sobre 300 = 30),
    # fracao 2 = 700 (tudo Serenus).
    checa('bruto esperado = 1000', abs(visao['bruto_esperado'] - 1000.0) < 0.01,
          visao['bruto_esperado'])
    checa('bruto do gestor = 300', abs(visao['bruto_gestor'] - 300.0) < 0.01,
          visao['bruto_gestor'])
    checa('retencao = 30', abs(visao['retencao_total'] - 30.0) < 0.01, visao['retencao_total'])
    checa('liquido pro PIX = 270', abs(visao['liquido_pix_gestor'] - 270.0) < 0.01,
          visao['liquido_pix_gestor'])
    checa('saldo Serenus = 700', abs(visao['saldo_serenus'] - 700.0) < 0.01,
          visao['saldo_serenus'])
    checa('nada de entrada confirmada ainda', abs(visao['entrada_confirmada']) < 0.01,
          visao['entrada_confirmada'])

    c = cliente()
    corpos = {}
    for rota in ('/financeiro', '/fluxo-caixa'):
        r = c.get(rota)
        checa(f'{rota} abre', r.status_code == 200, r.status_code)
        corpos[rota] = r.get_data(as_text=True)
    for rotulo in ('Bruto esperado', 'Apurado pela Affinity', 'Entrada confirmada',
                   'Bruto do gestor', 'Retenção', 'Líquido para PIX', 'Saldo Serenus'):
        for rota, corpo in corpos.items():
            checa(f'{rota} mostra "{rotulo}"', rotulo in corpo)
    # Os MESMOS numeros nas duas telas.
    for valor in ('1.000,00', '300,00', '270,00', '700,00'):
        checa(f'as duas telas mostram {valor}',
              all(valor in corpo for corpo in corpos.values()))


def teste_cor_de_saida():
    print('\n[unit] verde so pra entrada; obrigacao usa cor de saida')
    raiz = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    with open(os.path.join(raiz, 'templates/_painel_razao.html'), encoding='utf-8') as f:
        painel = f.read()
    for termo in ('Bruto do gestor', 'Retenção', 'Líquido para PIX'):
        pos = painel.find(termo)
        antes = painel[max(0, pos - 260):pos]
        checa(f'"{termo}" usa a classe de saida', 'pr-saida' in antes, antes[-60:])
    for termo in ('Entrada confirmada', 'Saldo Serenus'):
        pos = painel.find(termo)
        antes = painel[max(0, pos - 260):pos]
        checa(f'"{termo}" usa a classe de entrada', 'pr-entrada' in antes, antes[-60:])
    with open(os.path.join(raiz, 'templates/base.html'), encoding='utf-8') as f:
        base = f.read()
    checa('saida e vermelha, nao verde', '.pr-card.pr-saida .pr-v { color:var(--erro); }' in base)
    checa('entrada e verde', '.pr-card.pr-entrada .pr-v { color:var(--verde); }' in base)


def teste_exportacao_declara_escopo():
    print('\n[int] a exportacao diz qual recorte ela e')
    limpar()
    montar_venda_completa('Amil')
    c = cliente()
    c.post('/lancamento/salvar', json={'tipo': 'custo', 'descricao': 'TESTE FIN CUSTO',
                                       'valor': 100.0, 'data_emissao': '2026-08-05',
                                       'data_vencimento': '2026-08-10', 'num_parcelas': 1})
    r = c.get('/financeiro/exportar?mes=2026-08&secao=custos&formato=csv')
    corpo = r.get_data(as_text=True)
    checa('CSV de custos responde', r.status_code == 200, r.status_code)
    # O CSV comeca com BOM (pro Excel abrir com acento certo) e o csv.writer
    # poe aspas quando o texto tem ponto e virgula. Nada disso e conteudo.
    def primeira(t):
        return t.split('\n')[0].lstrip('\ufeff').strip('"')
    checa('declara a secao na primeira linha', primeira(corpo).startswith('Seção: custos'),
          primeira(corpo)[:80])
    checa('declara os filtros aplicados', 'Filtros:' in corpo)
    checa('avisa que a outra secao nao esta aqui', 'só a seção de custos e fixos' in corpo)
    checa('traz o total', 'TOTAL' in corpo)

    r2 = c.get('/financeiro/exportar?mes=2026-08&secao=comissao&formato=csv')
    corpo2 = r2.get_data(as_text=True)
    checa('CSV de comissao responde', r2.status_code == 200, r2.status_code)
    checa('declara a secao de comissao',
          primeira(corpo2).startswith('Seção: comissão'), primeira(corpo2)[:80])
    for rotulo in ('Bruto esperado', 'Entrada confirmada', 'Saldo Serenus'):
        checa(f'CSV de comissao traz "{rotulo}"', rotulo in corpo2)
    checa('os dois CSV sao diferentes', corpo != corpo2)

    rx = c.get('/financeiro/exportar?mes=2026-08&formato=xlsx')
    checa('XLSX responde', rx.status_code == 200, rx.status_code)
    try:
        import openpyxl, io
        wb = openpyxl.load_workbook(io.BytesIO(rx.get_data()))
        checa('XLSX tem as duas abas',
              wb.sheetnames == ['Custos e fixos', 'Comissao Affinity'], wb.sheetnames)
        checa('a aba de custos declara o escopo na primeira celula',
              str(wb['Custos e fixos']['A1'].value or '').startswith('Seção: custos'),
              wb['Custos e fixos']['A1'].value)
        checa('a aba de comissao declara o escopo',
              str(wb['Comissao Affinity']['A1'].value or '').startswith('Seção: comissão'),
              wb['Comissao Affinity']['A1'].value)
    except ImportError:
        print('    (openpyxl ausente — conteudo do XLSX nao conferido)')


def teste_permissoes():
    print('\n[unit] historico so por admin')
    consultor = A.app.test_client()
    with consultor.session_transaction() as s:
        s['user_id'] = 77
        s['perfil'] = 'consultor'
    r = consultor.get('/comissoes/regra-gestor/simular')
    checa('consultor nao simula', r.status_code in (302, 403), r.status_code)
    r = consultor.post('/comissoes/regra-gestor/aplicar-historico',
                       json={'propostas': [1], 'confirmacao': 'APLICAR'})
    checa('consultor nao aplica historico', r.status_code in (302, 403), r.status_code)


if __name__ == '__main__':
    print('=' * 66)
    print('HISTORICO E INTEGRACAO FINANCEIRO/FLUXO DE CAIXA — Fase 4')
    print('=' * 66)
    teste_vencimento_por_parcela()
    teste_politica_dia_31()
    teste_lancamento_parcelado_na_rota()
    teste_simulador_nao_escreve()
    teste_historico_protegido()
    teste_uma_fonte_para_as_duas_telas()
    teste_cor_de_saida()
    teste_exportacao_declara_escopo()
    teste_permissoes()
    print('\n' + '=' * 66)
    if falhas:
        print(f'FALHAS ({len(falhas)}):')
        for f in falhas:
            print('  - ' + f)
        sys.exit(1)
    print('Tudo passou.')
